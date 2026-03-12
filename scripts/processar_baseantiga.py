from __future__ import annotations

import csv
from pathlib import Path
from statistics import mean
from typing import Dict, List, Optional

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
INPUT_XLSX = BASE_DIR / 'baseantiga.xlsx'
OUT_CSV = BASE_DIR / 'data' / 'baseantiga_processada.csv'
OUT_RESUMO = BASE_DIR / 'data' / 'baseantiga_resumo.txt'

# Mesma metodologia do app atual (config.js)
PONTOS_Q1 = {1: 0, 2: 2, 3: 3, 4: 2, 5: 1}
PONTOS_Q2 = {1: 0, 2: 2, 3: 1}
PONTOS_Q5 = {1: 0, 2: 2, 3: 1}
PONTOS_Q6 = {1: 1, 2: 0, 3: 1}
PONTOS_DEFAULT = {1: 2, 2: 0, 3: 1}

GRUPOS = {
    'INPUT': [1],
    'RESIDUOS': [2],
    'OUTPUT': [3, 4, 5, 6],
    'VIDA': [7, 8, 9],
    'MONITORAMENTO': [10, 11, 12],
}
PESOS = {
    'INPUT': 0.25,
    'RESIDUOS': 0.20,
    'OUTPUT': 0.20,
    'VIDA': 0.20,
    'MONITORAMENTO': 0.15,
}

MAPEAMENTO = {
    1: 'MATERIA_PRIMA',
    2: 'RESIDUOS',
    3: 'DESMONTE',
    4: 'DESCARTE',
    5: 'RECUPERACAO',
    6: 'RECICLAGEM',
    7: 'DURABILIDADE',
    8: 'REPARAVEL',
    9: 'REAPROVEITAVEL',
    10: 'CICLO_ESTENDIDO',
    11: 'CICLO_RASTREADO',
    12: 'DOCUMENTACAO',
}


def norm_text(v: object) -> str:
    if v is None:
        return ''
    return ' '.join(str(v).strip().split())


def to_int_opt(v: object) -> Optional[int]:
    if v is None:
        return None
    s = str(v).strip()
    if s == '' or s.upper() in {'NULL', 'NAN', 'NONE'}:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def score_for_question(qid: int, answer: Optional[int]) -> int:
    if answer is None:
        return 0
    if qid == 1:
        return PONTOS_Q1.get(answer, 0)
    if qid == 2:
        return PONTOS_Q2.get(answer, 0)
    if qid == 5:
        return PONTOS_Q5.get(answer, 0)
    if qid == 6:
        return PONTOS_Q6.get(answer, 0)
    return PONTOS_DEFAULT.get(answer, 0)


def max_for_question(qid: int) -> int:
    if qid == 1:
        return max(PONTOS_Q1.values())
    if qid == 2:
        return max(PONTOS_Q2.values())
    if qid == 5:
        return max(PONTOS_Q5.values())
    if qid == 6:
        return max(PONTOS_Q6.values())
    return max(PONTOS_DEFAULT.values())


def processar() -> None:
    wb = load_workbook(INPUT_XLSX, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col = {h: i for i, h in enumerate(header)}

    out_rows: List[Dict[str, object]] = []
    total_validos = 0

    for row in rows:
        if row is None or all(v is None or norm_text(v) == '' for v in row):
            continue

        respostas: Dict[int, Optional[int]] = {}
        for qid, nome_col in MAPEAMENTO.items():
            respostas[qid] = to_int_opt(row[col[nome_col]])

        # Considera "registro válido" se tiver respostas numéricas das 12 questões
        valido = all(respostas[q] is not None for q in range(1, 13))
        if valido:
            total_validos += 1

        soma_pontos = sum(score_for_question(qid, respostas[qid]) for qid in range(1, 13))
        total_possivel = sum(max_for_question(qid) for qid in range(1, 13))
        igc = round((soma_pontos / total_possivel) * 100, 2) if total_possivel else 0.0

        grupos_perc: Dict[str, int] = {}
        soma_pond = 0.0
        soma_pesos = 0.0
        for grupo, qids in GRUPOS.items():
            pts = sum(score_for_question(qid, respostas[qid]) for qid in qids)
            max_g = sum(max_for_question(qid) for qid in qids)
            perc = round((pts / max_g) * 100) if max_g else 0
            grupos_perc[grupo] = int(perc)
            soma_pond += perc * PESOS[grupo]
            soma_pesos += PESOS[grupo]

        ime = round(soma_pond / soma_pesos, 2) if soma_pesos else igc

        out_rows.append({
            'ID_EMPRESA': row[col['ID_EMPRESA']],
            'NOME_EMPRESA': norm_text(row[col['NOME_EMPRESA']]),
            'CIDADE': norm_text(row[col['CIDADE']]).upper(),
            'ESTADO': norm_text(row[col['ESTADO']]).upper(),
            'SETOR': norm_text(row[col['SETOR']]).title(),
            'PRODUTO': norm_text(row[col['PRODUTO']]),
            'DATA_CADASTRO': row[col['DATA_CADASTRO']],
            'Q1_MATERIA_PRIMA': respostas[1],
            'Q2_RESIDUOS': respostas[2],
            'Q3_DESMONTE': respostas[3],
            'Q4_DESCARTE': respostas[4],
            'Q5_RECUPERACAO': respostas[5],
            'Q6_RECICLAGEM': respostas[6],
            'Q7_DURABILIDADE': respostas[7],
            'Q8_REPARAVEL': respostas[8],
            'Q9_REAPROVEITAVEL': respostas[9],
            'Q10_CICLO_ESTENDIDO': respostas[10],
            'Q11_CICLO_RASTREADO': respostas[11],
            'Q12_DOCUMENTACAO': respostas[12],
            'SOMA_CALCULADA': soma_pontos,
            'INDICE_GLOBAL_CALCULADO': igc,
            'MATURIDADE_CALCULADA': ime,
            'TOPICO_INPUT_%': grupos_perc['INPUT'],
            'TOPICO_RESIDUOS_%': grupos_perc['RESIDUOS'],
            'TOPICO_OUTPUT_%': grupos_perc['OUTPUT'],
            'TOPICO_VIDA_%': grupos_perc['VIDA'],
            'TOPICO_MONITORAMENTO_%': grupos_perc['MONITORAMENTO'],
            'REGISTRO_COMPLETO_12Q': 'SIM' if valido else 'NAO',
        })

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open('w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=list(out_rows[0].keys()))
        writer.writeheader()
        writer.writerows(out_rows)

    medias_igc = [float(r['INDICE_GLOBAL_CALCULADO']) for r in out_rows]
    medias_ime = [float(r['MATURIDADE_CALCULADA']) for r in out_rows]
    resumo = [
        f"Registros totais: {len(out_rows)}",
        f"Registros completos (12 questões): {total_validos}",
        f"Média IGC calculada: {mean(medias_igc):.2f}",
        f"Média IME calculada: {mean(medias_ime):.2f}",
        f"CSV gerado: {OUT_CSV}",
    ]
    OUT_RESUMO.write_text('\n'.join(resumo), encoding='utf-8')
    print('\n'.join(resumo))


if __name__ == '__main__':
    processar()
